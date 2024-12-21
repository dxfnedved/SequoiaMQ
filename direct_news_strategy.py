import os
from datetime import datetime
from news_sources.news_aggregator import NewsAggregator
from llm_interface import LLMInterface
from logger_manager import LoggerManager
from colorama import init, Fore, Style
import traceback
import logging
from openpyxl import Workbook
import pandas as pd

# 初始化colorama
init(autoreset=True)

class DirectNewsStrategy:
    """独立的新闻策略分析工具"""
    
    def __init__(self):
        """初始化新闻策略分析器"""
        self.logger = logging.getLogger(__name__)
        self.llm_interface = LLMInterface()
        self.news_aggregator = NewsAggregator(self.logger)
        self.focus_industries = [
            "新能源", "芯片", "人工智能", "医药", "新材料",
            "军工", "汽车", "消费", "金融", "地产"
        ]
        
    def run(self):
        """运行新闻策略分析"""
        try:
            print(f"\n{Fore.CYAN}开始执行新闻策略分析...{Style.RESET_ALL}")
            
            # 执行分析
            analysis_result = self._run_analysis()
            
            if analysis_result:
                # 生成报告
                print(f"\n{Fore.CYAN}正在生成分析报告...{Style.RESET_ALL}")
                self._generate_reports(analysis_result)
            else:
                print(f"{Fore.RED}分析结果为空，无法生成报告{Style.RESET_ALL}")
                
        except Exception as e:
            print(f"{Fore.RED}执行新闻策略分析失败: {str(e)}{Style.RESET_ALL}")
            self.logger.error(f"执行新闻策略分析失败: {str(e)}")
            self.logger.error(traceback.format_exc())
            
    def _run_analysis(self):
        """执行新闻分析流程"""
        try:
            print(f"{Fore.CYAN}开始获取新闻数据...{Style.RESET_ALL}")
            
            # 获取新闻数据
            news_data = self.news_aggregator.get_news()
            
            if news_data is None or news_data.empty:
                print(f"{Fore.RED}未获取到新闻数据{Style.RESET_ALL}")
                raise Exception("新闻数据获取失败")
                
            print(f"{Fore.GREEN}成功获取 {len(news_data)} 条新闻{Style.RESET_ALL}")
            
            # 分析新闻
            analysis_result = self._analyze_news(news_data)
            if not analysis_result:
                raise Exception("新闻分析失败")
            
            return analysis_result
            
        except Exception as e:
            print(f"{Fore.RED}新闻分析流程失败: {str(e)}{Style.RESET_ALL}")
            self.logger.error(f"新闻分析流程失败: {str(e)}")
            self.logger.error(traceback.format_exc())
            raise
            
    def _analyze_news(self, news_data):
        """使用COT方法通过多轮对话分析新闻数据"""
        try:
            print(f"{Fore.CYAN}开始多轮对话分析新闻数据...{Style.RESET_ALL}")
            
            # 第一轮：新闻分类和初步分析
            classification_prompt = self._get_classification_prompt(news_data)
            classification_result = self.llm_interface.analyze_with_cot(
                prompt=classification_prompt,
                context="请对新闻进行分类和初步分析，思考每条新闻的影响领域和重要程度。"
            )
            
            if not classification_result:
                raise Exception("新闻分类分析失败")
                
            print(f"{Fore.GREEN}完成新闻分类分析{Style.RESET_ALL}")
            
            # 第二轮：市场情绪分析
            sentiment_prompt = self._get_sentiment_prompt(news_data, classification_result)
            sentiment_result = self.llm_interface.analyze_with_cot(
                prompt=sentiment_prompt,
                context="基于新闻分类结果，分析市场情绪和投资者心理。"
            )
            
            if not sentiment_result:
                raise Exception("市场情绪分析失败")
                
            print(f"{Fore.GREEN}完成市场情绪分析{Style.RESET_ALL}")
            
            # 第三轮：行业影响分析
            industry_prompt = self._get_industry_prompt(classification_result, sentiment_result)
            industry_result = self.llm_interface.analyze_with_cot(
                prompt=industry_prompt,
                context="结合新闻分类和市场情绪，分析各行业的发展趋势和投资机会。"
            )
            
            if not industry_result:
                raise Exception("行业影响分析失败")
                
            print(f"{Fore.GREEN}完成行业影响分析{Style.RESET_ALL}")
            
            # 第四轮：风险评估
            risk_prompt = self._get_risk_prompt(classification_result, sentiment_result, industry_result)
            risk_result = self.llm_interface.analyze_with_cot(
                prompt=risk_prompt,
                context="基于前面的分析，评估当前市场的主要风险因素。"
            )
            
            if not risk_result:
                raise Exception("风险评估失败")
                
            print(f"{Fore.GREEN}完成风险评估{Style.RESET_ALL}")
            
            # 最终轮：生成投资策略
            strategy_prompt = self._get_strategy_prompt(
                classification_result,
                sentiment_result,
                industry_result,
                risk_result
            )
            final_result = self.llm_interface.analyze_with_cot(
                prompt=strategy_prompt,
                context="综合所有分析结果，生成详细的投资策略建议。"
            )
            
            if not final_result:
                raise Exception("投资策略生成失败")
                
            print(f"{Fore.GREEN}完成投资策略生成{Style.RESET_ALL}")
            
            return final_result
            
        except Exception as e:
            print(f"{Fore.RED}新闻分析失败: {str(e)}{Style.RESET_ALL}")
            self.logger.error(f"新闻分析失败: {str(e)}")
            self.logger.error(traceback.format_exc())
            raise
            
    def _get_classification_prompt(self, news_data):
        """生成新闻分类提示词"""
        news_summary = self._prepare_news_summary(news_data)
        return f"""
        作为金融分析师，请对以下新闻进行分类和初步分析：
        
        {news_summary}
        
        请按以下步骤思考：
        1. 将新闻按主题分类（如政策、行业、公司、市场等）
        2. 评估每类新闻的重要程度
        3. 分析每类新闻的潜在影响
        4. 识别关键事件和转折点
        
        输出格式：
        {{
            "news_classification": [
                {{
                    "category": "类别",
                    "importance": "重要程度",
                    "key_points": ["要点1", "要点2"],
                    "potential_impact": "潜在影响"
                }}
            ],
            "key_events": ["事件1", "事件2"],
            "initial_thoughts": "初步想法"
        }}
        """
        
    def _get_sentiment_prompt(self, news_data, classification_result):
        """生成市场情绪分析提示词"""
        return f"""
        基于前面的新闻分类结果：
        {classification_result}
        
        请分析市场情绪，思考以下方面：
        1. 整体市场情绪（乐观/中性/悲观）
        2. 各类投资者的心理状态
        3. 情绪变化的关键因素
        4. 可能的市场反应
        
        输出格式：
        {{
            "market_sentiment": {{
                "overall": "整体情绪",
                "retail_investors": "散户情绪",
                "institutional_investors": "机构情绪",
                "key_factors": ["因素1", "因素2"]
            }},
            "sentiment_analysis": "详细分析",
            "potential_reactions": ["反应1", "反应2"]
        }}
        """
        
    def _get_industry_prompt(self, classification_result, sentiment_result):
        """生成行业分析提示词"""
        return f"""
        基于新闻分类和市场情绪分析结果：
        分类结果：{classification_result}
        情绪分析：{sentiment_result}
        
        请分析以下行业的发展趋势：
        {self.focus_industries}
        
        思考步骤：
        1. 评估新闻对各行业的影响
        2. 分析行业发展机遇和挑战
        3. 识别潜在的投资机会
        4. 预测行业变化趋势
        
        输出格式：
        {{
            "industry_analysis": [
                {{
                    "industry": "行业名称",
                    "trend": "发展趋势",
                    "opportunities": ["机会1", "机会2"],
                    "challenges": ["挑战1", "挑战2"],
                    "key_companies": ["公司1", "公司2"]
                }}
            ],
            "cross_industry_impacts": ["影响1", "影响2"],
            "investment_opportunities": ["机会1", "机会2"]
        }}
        """
        
    def _get_risk_prompt(self, classification_result, sentiment_result, industry_result):
        """生成风险评估提示词"""
        return f"""
        基于之前的分析结果：
        分类结果：{classification_result}
        情绪分析：{sentiment_result}
        行业分析：{industry_result}
        
        请评估当前市场风险，思考：
        1. 主要风险因素
        2. 风险影响程度
        3. 风险发生概率
        4. 应对策略
        
        输出格式：
        {{
            "risk_assessment": [
                {{
                    "risk_type": "风险类型",
                    "probability": "发生概率",
                    "impact": "影响程度",
                    "mitigation": "应对策略"
                }}
            ],
            "overall_risk_level": "整体风险水平",
            "key_warnings": ["警示1", "警示2"]
        }}
        """
        
    def _get_strategy_prompt(self, classification_result, sentiment_result, industry_result, risk_result):
        """生成投资策略提示词"""
        return f"""
        基于所有分析结果：
        分类结果：{classification_result}
        情绪分析：{sentiment_result}
        行业分析：{industry_result}
        风险评估：{risk_result}
        
        请生成详细的投资策略建议，思考：
        1. 市场趋势判断
        2. 行业配置建议
        3. 具体操作策略
        4. 风险控制措施
        
        输出格式：
        {{
            "market_trend": {{
                "direction": "上涨/震荡/下跌",
                "analysis": "市场分析",
                "confidence": "高/中/低",
                "volume_analysis": "成交量分析",
                "northbound_flow": "北向资金分析"
            }},
            "industry_analysis": [
                {{
                    "industry": "行业名称",
                    "trend": "上涨/震荡/下跌",
                    "analysis": "行业分析",
                    "hot_topics": ["热点1", "热点2"],
                    "key_companies": ["公司1", "公司2"],
                    "risk_factors": ["风险1", "风险2"]
                }}
            ],
            "risk_warnings": [
                {{
                    "type": "风险类型",
                    "description": "风险描述",
                    "impact": "影响程度",
                    "suggestions": "建议"
                }}
            ],
            "investment_advice": {{
                "short_term": "短期建议",
                "mid_term": "中期建议",
                "focus_sectors": ["行业1", "行业2"],
                "operation_strategy": "操作策略",
                "risk_control": "风险控制"
            }}
        }}
        """
        
    def _prepare_news_summary(self, news_data):
        """准备新闻摘要"""
        try:
            if news_data is None or news_data.empty:
                return "无新闻数据"
                
            news_items = []
            for _, row in news_data.iterrows():
                title = str(row.get('title', '')).strip()
                content = str(row.get('content', '')).strip()
                time = str(row.get('time', '')).strip()
                source = str(row.get('source', '')).strip()
                
                if not title or not content:
                    continue
                    
                # 限制内容长度
                content = content[:500] + '...' if len(content) > 500 else content
                
                news_item = (
                    f"来源：{source}\n"
                    f"时间：{time}\n"
                    f"标题：{title}\n"
                    f"内容：{content}\n"
                )
                news_items.append(news_item)
                
            return "\n---\n".join(news_items)
            
        except Exception as e:
            self.logger.error(f"准备新闻摘要失败: {str(e)}")
            return "新闻摘要准备失败"
        
    def _generate_reports(self, analysis_result):
        """生成分析报告"""
        try:
            # 创建报告目录
            report_dir = os.path.join('summary', 'news_analysis')
            os.makedirs(report_dir, exist_ok=True)
            
            # 生成报告文件名
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_file = os.path.join(report_dir, f'news_analysis_{timestamp}.txt')
            excel_file = os.path.join(report_dir, f'news_analysis_{timestamp}.xlsx')
            
            # 生成文本报告
            self._generate_text_report(analysis_result, report_file)
            
            # 生成Excel报告
            self._generate_excel_report(analysis_result, excel_file)
            
            print(f"\n{Fore.GREEN}分析报告已生成：")
            print(f"文本报告：{report_file}")
            print(f"Excel报告：{excel_file}{Style.RESET_ALL}")
            
        except Exception as e:
            print(f"{Fore.RED}生成报告失败: {str(e)}{Style.RESET_ALL}")
            self.logger.error(f"生成报告失败: {str(e)}")
            
    def _generate_text_report(self, analysis_result, report_file):
        """生成文本格式的分析报告"""
        try:
            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("=== 新闻分析报告 ===\n\n")
                f.write(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                
                # 市场趋势
                market_trend = analysis_result.get('market_trend', {})
                f.write("== 市场趋势 ==\n")
                f.write(f"大盘方向: {market_trend.get('direction', '未知')}\n")
                f.write(f"分析结果: {market_trend.get('analysis', '未知')}\n")
                f.write(f"置信度: {market_trend.get('confidence', '未知')}\n")
                f.write(f"成交量分析: {market_trend.get('volume_analysis', '未知')}\n")
                f.write(f"北向资金: {market_trend.get('northbound_flow', '未知')}\n\n")
                
                # 行业分析
                f.write("== 行业分析 ==\n")
                for industry in analysis_result.get('industry_analysis', []):
                    f.write(f"\n= {industry.get('industry', '未知行业')} =\n")
                    f.write(f"走势: {industry.get('trend', '未知')}\n")
                    f.write(f"分析: {industry.get('analysis', '未知')}\n")
                    f.write(f"热点: {', '.join(industry.get('hot_topics', ['未知']))}\n")
                    f.write(f"关键公司: {', '.join(industry.get('key_companies', ['未知']))}\n")
                    f.write(f"风险因素: {', '.join(industry.get('risk_factors', ['未知']))}\n")
                
                # 风险提示
                f.write("\n== 风险提示 ==\n")
                for risk in analysis_result.get('risk_warnings', []):
                    f.write(f"\n类型: {risk.get('type', '未知')}\n")
                    f.write(f"描述: {risk.get('description', '未知')}\n")
                    f.write(f"影响: {risk.get('impact', '未知')}\n")
                    f.write(f"建议: {risk.get('suggestions', '未知')}\n")
                
                # 投资建议
                investment = analysis_result.get('investment_advice', {})
                f.write("\n== 投资建议 ==\n")
                f.write(f"短期建议: {investment.get('short_term', '未知')}\n")
                f.write(f"中期建议: {investment.get('mid_term', '未知')}\n")
                f.write(f"关注板块: {', '.join(investment.get('focus_sectors', ['未知']))}\n")
                f.write(f"操作策略: {investment.get('operation_strategy', '未知')}\n")
                f.write(f"风险控制: {investment.get('risk_control', '未知')}\n")
                
            print(f"{Fore.GREEN}文本报告生成完成{Style.RESET_ALL}")
            
        except Exception as e:
            print(f"{Fore.RED}生成文本报告失败: {str(e)}{Style.RESET_ALL}")
            self.logger.error(f"生成文本报告失败: {str(e)}")
            
    def _generate_excel_report(self, analysis_result, excel_file):
        """生成Excel格式的分析报告"""
        try:
            # 创建工作簿
            wb = Workbook()
            
            # 市场趋势sheet
            ws_market = wb.active
            ws_market.title = "市场趋势"
            market_trend = analysis_result.get('market_trend', {})
            ws_market.append(['指标', '内容'])
            ws_market.append(['大盘方向', market_trend.get('direction', '未知')])
            ws_market.append(['分析结果', market_trend.get('analysis', '未知')])
            ws_market.append(['置信度', market_trend.get('confidence', '未知')])
            ws_market.append(['成交量分析', market_trend.get('volume_analysis', '未知')])
            ws_market.append(['北向资金', market_trend.get('northbound_flow', '未知')])
            
            # 行业分析sheet
            ws_industry = wb.create_sheet("行业分析")
            ws_industry.append(['行业', '走势', '分析', '热点', '关键公司', '风险因素'])
            for industry in analysis_result.get('industry_analysis', []):
                ws_industry.append([
                    industry.get('industry', '未知'),
                    industry.get('trend', '未知'),
                    industry.get('analysis', '未知'),
                    '\n'.join(industry.get('hot_topics', ['未知'])),
                    '\n'.join(industry.get('key_companies', ['未知'])),
                    '\n'.join(industry.get('risk_factors', ['未知']))
                ])
            
            # 风险提示sheet
            ws_risk = wb.create_sheet("风险提示")
            ws_risk.append(['类型', '描述', '影响', '建议'])
            for risk in analysis_result.get('risk_warnings', []):
                ws_risk.append([
                    risk.get('type', '未知'),
                    risk.get('description', '未知'),
                    risk.get('impact', '未知'),
                    risk.get('suggestions', '未知')
                ])
            
            # 投资建议sheet
            ws_investment = wb.create_sheet("投资建议")
            investment = analysis_result.get('investment_advice', {})
            ws_investment.append(['指标', '内容'])
            ws_investment.append(['短期建议', investment.get('short_term', '未知')])
            ws_investment.append(['中期建议', investment.get('mid_term', '未知')])
            ws_investment.append(['关注板块', '\n'.join(investment.get('focus_sectors', ['未知']))])
            ws_investment.append(['操作策略', investment.get('operation_strategy', '未知')])
            ws_investment.append(['风险控制', investment.get('risk_control', '未知')])
            
            # 保存Excel文件
            wb.save(excel_file)
            print(f"{Fore.GREEN}Excel报告生成完成{Style.RESET_ALL}")
            
        except Exception as e:
            print(f"{Fore.RED}生成Excel报告失败: {str(e)}{Style.RESET_ALL}")
            self.logger.error(f"生成Excel报告失败: {str(e)}")

def main():
    """主函数"""
    strategy = DirectNewsStrategy()
    strategy.run()

if __name__ == '__main__':
    try:
        # 设置日志
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('logs/news_strategy.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        # 创建必要的目录
        os.makedirs('logs', exist_ok=True)
        os.makedirs('summary/news_analysis', exist_ok=True)
        
        main()
        
    except Exception as e:
        print(f"{Fore.RED}程序执行失败: {str(e)}{Style.RESET_ALL}")
        logging.error(f"程序执行失败: {str(e)}")
        logging.error(traceback.format_exc()) 